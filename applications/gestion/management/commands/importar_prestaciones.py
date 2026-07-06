import os

from openpyxl import load_workbook

from django.core.management.base import BaseCommand
from django.db import transaction
from django.contrib.auth import get_user_model

from applications.entidades.models import Prestacion


def limpiar_texto(valor):
    if valor is None:
        return None

    valor = str(valor).strip()
    return valor if valor else None


def limpiar_nombre(valor):
    valor = limpiar_texto(valor)
    return valor.title() if valor else ""


class Command(BaseCommand):
    help = "Importa prestaciones desde un archivo Excel"

    def add_arguments(self, parser):
        parser.add_argument("archivo", type=str)
        parser.add_argument("--dry-run", action="store_true")
        parser.add_argument("--confirm", action="store_true")
        parser.add_argument("--usuario", type=str)

    def handle(self, *args, **options):
        archivo = options["archivo"]
        dry_run = options["dry_run"]
        confirm = options["confirm"]
        usuario_username = options["usuario"]

        if not dry_run and not confirm:
            self.stdout.write(
                self.style.ERROR(
                    "Usá --dry-run para probar o --confirm para importar."
                )
            )
            return

        if not os.path.exists(archivo):
            self.stdout.write(
                self.style.ERROR(f"No existe el archivo: {archivo}")
            )
            return

        usuario_migracion = None

        if usuario_username:
            User = get_user_model()

            try:
                usuario_migracion = User.objects.get(username=usuario_username)
            except User.DoesNotExist:
                self.stdout.write(
                    self.style.ERROR(
                        f"No existe el usuario: {usuario_username}"
                    )
                )
                return

        self.stdout.write("Leyendo Excel...")

        wb = load_workbook(archivo, data_only=True)
        ws = wb.active

        creados = 0
        actualizados = 0
        omitidos = 0
        errores = 0
        procesados = 0

        try:
            with transaction.atomic():

                # Los datos comienzan en la fila 4 (la fila 3 contiene los encabezados)
                for fila in ws.iter_rows(min_row=4, values_only=True):

                    codigo = limpiar_texto(fila[0])
                    nombre = limpiar_nombre(fila[1])

                    if not codigo or not nombre:
                        omitidos += 1
                        continue

                    procesados += 1

                    try:

                        defaults = {
                            "nombre": nombre,
                        }

                        if usuario_migracion:
                            defaults["user_updated"] = usuario_migracion

                        prestacion = Prestacion.objects.filter(
                            codigo=str(codigo)
                        ).first()

                        if prestacion:

                            for campo, valor in defaults.items():
                                setattr(prestacion, campo, valor)

                            prestacion.save()
                            actualizados += 1

                        else:

                            if usuario_migracion:
                                defaults["user_made"] = usuario_migracion

                            Prestacion.objects.create(
                                codigo=str(codigo),
                                **defaults
                            )

                            creados += 1

                    except Exception as e:
                        errores += 1
                        self.stdout.write(
                            self.style.ERROR(
                                f"Error en código {codigo}: {e}"
                            )
                        )

                if dry_run:
                    raise Exception("DRY RUN")

        except Exception:
            if dry_run:
                self.stdout.write(
                    self.style.WARNING(
                        "Dry-run finalizado. No se guardó nada."
                    )
                )
            else:
                raise

        self.stdout.write(
            self.style.SUCCESS(
                f"""
        Importación finalizada

        Procesados : {procesados}
        Creados    : {creados}
        Actualizados: {actualizados}
        Omitidos   : {omitidos}
        Errores    : {errores}
        """
                    )
        )